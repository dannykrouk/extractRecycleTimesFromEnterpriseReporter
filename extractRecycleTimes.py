# This script will extract the recycle time information the GIS Enterprise Reporter "admin Excel" report (assuming the *svc sheets are in their default sort order)
# The outputs will be one .csv per ArcGIS Server Site in the admin Excel, placed in the same directory as the Excel file
# "c:\Program Files\ArcGIS\Server\framework\runtime\ArcGIS\bin\Python\Scripts\propy.bat" extractRecycleTimes.py -f C:\temp\reporterAdminExcel.xlsx 
import sys 
import os
import csv
import argparse

import openpyxl
from openpyxl import Workbook


def main(argv=None):
    
    global gOutputDir
    global gAdminExcel
    print('### BEGINNING ###')
    
    # get parameters
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file')
    args = parser.parse_args()
    
    gAdminExcel = args.file
    gOutputDir = os.path.dirname(gAdminExcel)
    
    # do work
    processExcel()
    
    print('### PROCESS COMPLETE ###')


# find all "svc" sheets and send each for processing
def processExcel():
    try:
        wb = openpyxl.load_workbook(gAdminExcel)
        sheetNames = wb.sheetnames
        for shName in sheetNames:
            if shName.endswith('svc'):
                processSheet(wb, shName)
    except Exception as e:
        print(e)        

# make a ServiceInfo object for each service in the sheet 
def processSheet(wb, sheetName):
    ws = wb[sheetName]
    
    siProperties = None # si class properties
    dictServiceInfos = {} # serviceName (string), ServiceInfo (object)

    # populate the dictionary
    si = None
    for row in ws.iter_rows():
        
        # skip the row header
        if row[1].value != 'urlPath':
            serviceName = row[1].value
            
            if si == None:
                # initialize si
                si = ServiceInfo(sheetName, serviceName)
            else:
                # populate si 
                if row[2].value == 'recycleInterval':
                    si.recycleInterval = str(row[3].value).strip('"')
                if row[2].value == 'recycleStartTime':
                    si.recycleStartTime = str(row[3].value).strip('"')
                if row[2].value == 'provider':
                    si.provider = str(row[3].value).strip('"')
                if row[2].value == 'maxStartupTime':
                    si.maxStartupTime = str(row[3].value).strip('"')
                if si.recycleInterval != '' and si.recycleStartTime != '' and si.provider != '' and si.maxStartupTime != '':
                    # add to dictionary and re-initialize si
                    si.serviceName = serviceName 
                    dictServiceInfos.update({serviceName:si})
                    if siProperties == None:
                        # create a list of the si class' properties (if we did not already)
                        props = si.__dict__
                        siProperties = list(props)
                    si = None

    # get si property names for csv header (get from first object)
    firstKey = next(iter(dictServiceInfos))
    firstValueSi = dictServiceInfos[firstKey]
    props = firstValueSi.__dict__
    propKeys = list(props)

    # write to csv
    fileName = os.path.join(gOutputDir, sheetName + ".csv")
    print ('Writing to: ' + fileName)

    with open (fileName, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(propKeys)
        #print('Wrote headers: ' + str(propKeys))
        i = 0
        for key, value in dictServiceInfos.items():
            si = value;
            propValues = list(vars(si).values())
            writer.writerow(propValues)
            i += 1
            #print('Wrote values: ' + str(propValues))
        print ('Wrote: ' + str(i) + ' records to csv')
   
class ServiceInfo:
    
    #ctor
    def __init__(self, sheetName, serviceName, provider, recycleInterval, recycleStartTime, maxStartupTime):
        self.sheetName = sheetName
        self.serviceName = serviceName
        self.provider = provider
        self.recycleInterval = recycleInterval
        self.recycleStartTime = recycleStartTime
        self.maxStartupTime = maxStartupTime

    def __init__(self, sheetName, serviceName):
        self.sheetName = sheetName
        self.serviceName = serviceName
        self.provider = ''
        self.recycleInterval = ''
        self.recycleStartTime = ''
        self.maxStartupTime = ''
    
    # print support
    def __str__(self):
        return f"ServiceInfo(sheetName={self.sheetName}, serviceName={self.serviceName}, provider={self.provider}, recycleInterval={self.recycleInterval}, recycleStartTime={self.recycleStartTime}, maxStartupTime={self.maxStartupTime} )"
      
if __name__ == "__main__":
    sys.exit(main(sys.argv))